from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic import ListView, DetailView
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Q
from django.contrib.admin.views.decorators import staff_member_required
from django.utils.decorators import method_decorator
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from datetime import datetime
from .models import Website, BasicAnalysis, SEOIssue, TranslitResult
from .seo_parser import SEOParser
from .translit_parser import TranslitParser


class WebsiteListView(ListView):
    """Список всех веб-сайтов"""
    model = Website
    template_name = 'tools/website_list.html'
    context_object_name = 'websites'
    paginate_by = 10
    
    def get_queryset(self):
        queryset = Website.objects.all()
        
        # Фильтр по конкурентам
        competitor_filter = self.request.GET.get('competitor')
        if competitor_filter == 'true':
            queryset = queryset.filter(is_competitor=True)
        elif competitor_filter == 'false':
            queryset = queryset.filter(is_competitor=False)
        
        # Поиск
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(url__icontains=search) |
                Q(description__icontains=search)
            )
        
        # Сортировка
        sort_by = self.request.GET.get('sort', '-created_at')
        if sort_by in ['name', '-name', 'url', '-url', 'created_at', '-created_at']:
            queryset = queryset.order_by(sort_by)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Статистика
        context['stats'] = {
            'total_websites': Website.objects.count(),
            'competitors': Website.objects.filter(is_competitor=True).count(),
            'total_analyses': BasicAnalysis.objects.count(),
        }
        
        return context


class WebsiteDetailView(DetailView):
    """Детальная страница веб-сайта"""
    model = Website
    template_name = 'tools/website_detail.html'
    context_object_name = 'website'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['analyses'] = self.object.analyses.all().order_by('-created_at')
        return context


class BasicAnalysisListView(ListView):
    """Список всех базовых анализов"""
    model = BasicAnalysis
    template_name = 'tools/analysis_list.html'
    context_object_name = 'analyses'
    paginate_by = 12
    
    def get_queryset(self):
        queryset = BasicAnalysis.objects.select_related('website')
        
        # Фильтр по сайту
        website_id = self.request.GET.get('website')
        if website_id:
            queryset = queryset.filter(website_id=website_id)
        
        # Фильтр по конкурентам
        competitor_filter = self.request.GET.get('competitor')
        if competitor_filter == 'true':
            queryset = queryset.filter(website__is_competitor=True)
        elif competitor_filter == 'false':
            queryset = queryset.filter(website__is_competitor=False)
        
        
        # Поиск
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(page_url__icontains=search) |
                Q(page_title__icontains=search) |
                Q(meta_description__icontains=search) |
                Q(website__name__icontains=search)
            )
        
        # Сортировка
        sort_by = self.request.GET.get('sort', '-created_at')
        if sort_by in ['page_title', '-page_title', 'created_at', '-created_at', 'response_time', '-response_time']:
            queryset = queryset.order_by(sort_by)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['websites'] = Website.objects.all()
        
        # Статистика
        context['stats'] = {
            'total_analyses': BasicAnalysis.objects.count(),
            'total_websites': Website.objects.count(),
            'competitors': Website.objects.filter(is_competitor=True).count(),
        }
        
        # Добавляем результаты транслита из сессии
        context['translit_results'] = self.request.session.get('translit_results')
        
        # Очищаем результаты транслита если запрошено
        if self.request.GET.get('clear_translit'):
            if 'translit_results' in self.request.session:
                del self.request.session['translit_results']
        
        return context


class BasicAnalysisDetailView(DetailView):
    """Детальная страница анализа"""
    model = BasicAnalysis
    template_name = 'tools/analysis_detail.html'
    context_object_name = 'analysis'


def analyze_website(request):
    """Анализ веб-сайта"""
    if request.method == 'POST':
        url = request.POST.get('url')
        website_name = request.POST.get('name', '')
        is_competitor = request.POST.get('is_competitor') == 'on'
        keywords = request.POST.get('keywords', '').strip()
        
        if not url:
            messages.error(request, 'Введите URL для анализа')
            return redirect('tools:website_list')
        
        try:
            # Создаем или получаем веб-сайт
            website, created = Website.objects.get_or_create(
                url=url,
                defaults={
                    'name': website_name or url,
                    'is_competitor': is_competitor
                }
            )
            
            # Парсим данные
            parser = SEOParser()
            keywords_list = [k.strip() for k in keywords.split('\n') if k.strip()] if keywords else None
            data = parser.parse_url(url, keywords_list)
            
            if 'error' in data:
                messages.error(request, f'Ошибка анализа: {data["error"]}')
                return redirect('tools:website_list')
            
            # Получаем robots.txt и sitemap.xml
            data['robots_txt'] = parser.get_robots_txt(url)
            data['sitemap_xml'] = parser.get_sitemap_xml(url)
            
            # Создаем анализ
            analysis = BasicAnalysis.objects.create(
                website=website,
                **data
            )
            
            # Создаем SEO проблемы
            for issue_data in data.get('seo_issues', []):
                SEOIssue.objects.create(
                    analysis=analysis,
                    **issue_data
                )
            
            messages.success(request, f'Анализ завершен! Найдено {data.get("word_count", 0)} слов, {data.get("total_links", 0)} ссылок.')
            return redirect('tools:analysis_detail', pk=analysis.pk)
            
        except Exception as e:
            messages.error(request, f'Ошибка при анализе: {str(e)}')
            return redirect('tools:website_list')
    
    return redirect('tools:website_list')


def export_to_excel(request):
    """Экспорт анализов в Excel"""
    # Проверяем, экспортируем ли отдельный анализ
    analysis_id = request.GET.get('analysis')
    if analysis_id:
        analyses = BasicAnalysis.objects.filter(pk=analysis_id).select_related('website')
    else:
        # Получаем отфильтрованные данные
        analyses = BasicAnalysis.objects.select_related('website')
        
        # Применяем фильтры
        website_id = request.GET.get('website')
        if website_id:
            analyses = analyses.filter(website_id=website_id)
        
        competitor_filter = request.GET.get('competitor')
        if competitor_filter == 'true':
            analyses = analyses.filter(website__is_competitor=True)
        elif competitor_filter == 'false':
            analyses = analyses.filter(website__is_competitor=False)
    
    search = request.GET.get('search')
    if search:
        analyses = analyses.filter(
            Q(page_url__icontains=search) |
            Q(page_title__icontains=search) |
            Q(meta_description__icontains=search) |
            Q(website__name__icontains=search)
        )
    
    # Создаем Excel файл
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SEO Анализ"
    
    # Заголовки
    headers = [
        'Сайт', 'URL страницы', 'Title', 'Meta Description', 'Meta Keywords',
        'Meta Robots', 'Canonical', 'H1', 'H2', 'H3', 'H4', 'H5', 'H6',
        'OG Title', 'OG Description', 'OG Image', 'OG Type',
        'Twitter Title', 'Twitter Description', 'Twitter Image', 'Twitter Card',
        'Время ответа', 'Размер страницы', 'HTTP статус', 'Количество слов',
        'Длина title', 'Длина description', 'Внутренние ссылки', 'Внешние ссылки',
        'Всего ссылок', 'Извлеченный текст', 'Дата анализа'
    ]
    
    # Стили
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Записываем заголовки
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Записываем данные
    for row, analysis in enumerate(analyses, 2):
        ws.cell(row=row, column=1, value=analysis.website.name)
        ws.cell(row=row, column=2, value=analysis.page_url)
        ws.cell(row=row, column=3, value=analysis.page_title)
        ws.cell(row=row, column=4, value=analysis.meta_description)
        ws.cell(row=row, column=5, value=analysis.meta_keywords)
        ws.cell(row=row, column=6, value=analysis.meta_robots)
        ws.cell(row=row, column=7, value=analysis.canonical_url)
        ws.cell(row=row, column=8, value=', '.join(analysis.h1_tags))
        ws.cell(row=row, column=9, value=', '.join(analysis.h2_tags))
        ws.cell(row=row, column=10, value=', '.join(analysis.h3_tags))
        ws.cell(row=row, column=11, value=', '.join(analysis.h4_tags))
        ws.cell(row=row, column=12, value=', '.join(analysis.h5_tags))
        ws.cell(row=row, column=13, value=', '.join(analysis.h6_tags))
        ws.cell(row=row, column=14, value=analysis.og_title)
        ws.cell(row=row, column=15, value=analysis.og_description)
        ws.cell(row=row, column=16, value=analysis.og_image)
        ws.cell(row=row, column=17, value=analysis.og_type)
        ws.cell(row=row, column=18, value=analysis.twitter_title)
        ws.cell(row=row, column=19, value=analysis.twitter_description)
        ws.cell(row=row, column=20, value=analysis.twitter_image)
        ws.cell(row=row, column=21, value=analysis.twitter_card)
        ws.cell(row=row, column=22, value=analysis.response_time)
        ws.cell(row=row, column=23, value=analysis.page_size)
        ws.cell(row=row, column=24, value=analysis.status_code)
        ws.cell(row=row, column=25, value=analysis.word_count)
        ws.cell(row=row, column=26, value=analysis.title_length)
        ws.cell(row=row, column=27, value=analysis.description_length)
        ws.cell(row=row, column=28, value=analysis.internal_links)
        ws.cell(row=row, column=29, value=analysis.external_links)
        ws.cell(row=row, column=30, value=analysis.total_links)
        ws.cell(row=row, column=31, value=analysis.extracted_text[:1000] + '...' if len(analysis.extracted_text) > 1000 else analysis.extracted_text)
        ws.cell(row=row, column=32, value=analysis.created_at.strftime('%d.%m.%Y %H:%M'))
    
    # Автоширина колонок
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем в память
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Возвращаем файл
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    # Определяем имя файла
    if analysis_id:
        analysis = analyses.first()
        filename = f"seo_analysis_{analysis.website.name}_{analysis.created_at.strftime('%Y%m%d_%H%M%S')}.xlsx"
    else:
        filename = f"seo_analyses_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


def download_robots_txt(request, analysis_id):
    """Скачивание robots.txt"""
    analysis = get_object_or_404(BasicAnalysis, pk=analysis_id)
    
    response = HttpResponse(analysis.robots_txt, content_type='text/plain')
    response['Content-Disposition'] = f'attachment; filename="robots_{analysis.website.name}.txt"'
    
    return response


def download_sitemap_xml(request, analysis_id):
    """Скачивание sitemap.xml"""
    analysis = get_object_or_404(BasicAnalysis, pk=analysis_id)
    
    response = HttpResponse(analysis.sitemap_xml, content_type='application/xml')
    response['Content-Disposition'] = f'attachment; filename="sitemap_{analysis.website.name}.xml"'
    
    return response


# ===== ТРАНСЛИТ =====

class TranslitListView(ListView):
    """Список результатов транслита"""
    model = TranslitResult
    template_name = 'tools/translit_list.html'
    context_object_name = 'results'
    paginate_by = 12
    
    def get_queryset(self):
        queryset = TranslitResult.objects.all()
        
        # Поиск
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(original_text__icontains=search) |
                Q(translit_text__icontains=search)
            )
        
        # Фильтр по типу
        is_url = self.request.GET.get('is_url')
        if is_url == 'true':
            queryset = queryset.filter(is_url=True)
        elif is_url == 'false':
            queryset = queryset.filter(is_url=False)
        
        return queryset.order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['stats'] = {
            'total_results': TranslitResult.objects.count(),
            'url_results': TranslitResult.objects.filter(is_url=True).count(),
            'text_results': TranslitResult.objects.filter(is_url=False).count(),
        }
        return context


def translit_form(request):
    """Форма транслитерации"""
    if request.method == 'POST':
        text = request.POST.get('text', '').strip()
        
        if not text:
            messages.error(request, 'Введите текст для транслитерации')
            return redirect('tools:analysis_list')
        
        try:
            parser = TranslitParser()
            # Всегда обрабатываем как обычный текст (не URL)
            results = parser.process_multiple_lines(text, False)
            
            # Объединяем все результаты в один список
            all_translits = []
            for result in results:
                # Убираем лишние пробелы
                clean_translit = result['translit'].strip()
                if clean_translit:  # Добавляем только непустые результаты
                    all_translits.append(clean_translit)
            
            # Сохраняем результаты в сессии для отображения
            request.session['translit_results'] = {
                'original_text': text.strip(),
                'translit_list': all_translits,
                'count': len(all_translits)
            }
            
            messages.success(request, f'Транслитерация завершена! Обработано {len(results)} строк.')
            return redirect('tools:analysis_list')
            
        except Exception as e:
            messages.error(request, f'Ошибка при транслитерации: {str(e)}')
            return redirect('tools:analysis_list')
    
    return redirect('tools:analysis_list')



