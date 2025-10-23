from django.shortcuts import render
from django.views.generic import ListView
from django.http import HttpResponse
from django.db.models import Sum, Count, Q
from django.db import models
from datetime import datetime, timedelta
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, Alignment
from .models import WorkReport


class ReportListView(ListView):
    model = WorkReport
    template_name = 'reports/report_list.html'
    context_object_name = 'reports'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = WorkReport.objects.select_related('project').all()
        
        # Фильтр по дате
        date_filter = self.request.GET.get('date')
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        
        if date_filter == 'today':
            today = timezone.now().date()
            queryset = queryset.filter(date=today)
        elif date_filter == 'week':
            week_ago = timezone.now().date() - timedelta(days=7)
            queryset = queryset.filter(date__gte=week_ago)
        elif date_filter == 'month':
            month_ago = timezone.now().date() - timedelta(days=30)
            queryset = queryset.filter(date__gte=month_ago)
        elif date_filter == 'year':
            year_ago = timezone.now().date() - timedelta(days=365)
            queryset = queryset.filter(date__gte=year_ago)
        
        # Фильтр по конкретным датам
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)
        
        # Фильтр по проекту
        project_filter = self.request.GET.get('project')
        if project_filter:
            queryset = queryset.filter(
                Q(project__name__icontains=project_filter) |
                Q(project_name__icontains=project_filter)
            )
        
        # Фильтр по выбранному проекту
        project_select = self.request.GET.get('project_select')
        if project_select:
            queryset = queryset.filter(project_id=project_select)
        
        # Фильтр по статусу проекта
        project_status = self.request.GET.get('project_status')
        if project_status:
            queryset = queryset.filter(project__status=project_status)
        
        # Поиск по названию проекта
        search_query = self.request.GET.get('search')
        if search_query:
            queryset = queryset.filter(
                Q(project__name__icontains=search_query) |
                Q(project_name__icontains=search_query) |
                Q(work_description__icontains=search_query)
            )
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Статистика
        today = timezone.now().date()
        week_ago = today - timedelta(days=7)
        month_ago = today - timedelta(days=30)
        
        context['stats'] = {
            'today_time': WorkReport.objects.filter(date=today).aggregate(
                total=Sum('time_spent')
            )['total'] or timedelta(),
            'week_time': WorkReport.objects.filter(date__gte=week_ago).aggregate(
                total=Sum('time_spent')
            )['total'] or timedelta(),
            'month_time': WorkReport.objects.filter(date__gte=month_ago).aggregate(
                total=Sum('time_spent')
            )['total'] or timedelta(),
            'total_reports': WorkReport.objects.count(),
            'projects_count': WorkReport.objects.values('project_name').distinct().count(),
        }
        
        # Фильтры
        context['date_choices'] = [
            ('', 'Все даты'),
            ('today', 'Сегодня'),
            ('week', 'За неделю'),
            ('month', 'За месяц'),
            ('year', 'За год'),
            ('custom', 'Выбрать период'),
        ]
        context['current_date_filter'] = self.request.GET.get('date', '')
        context['current_date_from'] = self.request.GET.get('date_from', '')
        context['current_date_to'] = self.request.GET.get('date_to', '')
        context['current_project_filter'] = self.request.GET.get('project', '')
        context['current_search'] = self.request.GET.get('search', '')
        context['selected_project'] = self.request.GET.get('project_select', '')
        context['selected_project_status'] = self.request.GET.get('project_status', '')
        
        # Список проектов для фильтра
        from projects.models import Project
        context['projects'] = Project.objects.all().order_by('name')
        
        # Статусы проектов
        context['project_status_choices'] = [
            ('', 'Все статусы'),
            ('in_progress', 'В работе'),
            ('ready', 'Готов'),
            ('archive', 'Архив'),
        ]
        
        return context


def export_to_excel(request):
    """Экспорт отчетов в Excel"""
    # Получаем фильтры точно так же, как в ReportListView
    date_filter = request.GET.get('date')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    project_filter = request.GET.get('project')
    project_select = request.GET.get('project_select')
    project_status = request.GET.get('project_status')
    search_query = request.GET.get('search')
    
    # Фильтруем данные точно так же, как в ReportListView
    queryset = WorkReport.objects.select_related('project').all()
    
    # Фильтр по дате (точно такая же логика, как в ReportListView)
    if date_filter == 'today':
        today = timezone.now().date()
        queryset = queryset.filter(date=today)
    elif date_filter == 'week':
        week_ago = timezone.now().date() - timedelta(days=7)
        queryset = queryset.filter(date__gte=week_ago)
    elif date_filter == 'month':
        month_ago = timezone.now().date() - timedelta(days=30)
        queryset = queryset.filter(date__gte=month_ago)
    elif date_filter == 'year':
        year_ago = timezone.now().date() - timedelta(days=365)
        queryset = queryset.filter(date__gte=year_ago)
    
    # Фильтр по конкретным датам
    if date_from:
        queryset = queryset.filter(date__gte=date_from)
    if date_to:
        queryset = queryset.filter(date__lte=date_to)
    
    # Фильтр по проекту (старый способ для совместимости)
    if project_filter:
        queryset = queryset.filter(
            Q(project__name__icontains=project_filter) |
            Q(project_name__icontains=project_filter)
        )
    
    # Фильтр по выбранному проекту
    if project_select:
        queryset = queryset.filter(project_id=project_select)
    
    # Фильтр по статусу проекта
    if project_status:
        queryset = queryset.filter(project__status=project_status)
    
    # Поиск
    if search_query:
        queryset = queryset.filter(
            Q(project__name__icontains=search_query) |
            Q(project_name__icontains=search_query) |
            Q(work_description__icontains=search_query)
        )
    
    
    # Создаем Excel файл
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчеты о работе"
    
    # Начинаем сразу с заголовков
    start_row = 1
    
    # Заголовки
    headers = ['Дата', 'Проект', 'Ссылка', 'Описание работы', 'Время']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Данные
    total_time = timedelta()
    data_start_row = start_row + 1
    for row, report in enumerate(queryset, data_start_row):
        ws.cell(row=row, column=1, value=report.date.strftime('%d.%m.%Y'))
        ws.cell(row=row, column=2, value=report.get_project_name())
        ws.cell(row=row, column=3, value=report.project_url or '')
        ws.cell(row=row, column=4, value=report.work_description)
        ws.cell(row=row, column=5, value=str(report.time_spent))
        total_time += report.time_spent
    
    # Добавляем итоговую строку
    if queryset.exists():
        total_row = len(queryset) + data_start_row + 1
        ws.cell(row=total_row, column=1, value="ИТОГО:")
        ws.cell(row=total_row, column=2, value=f"Записей: {len(queryset)}")
        ws.cell(row=total_row, column=5, value=str(total_time))
        
        # Стилизация итоговой строки
        for col in range(1, 6):
            cell = ws.cell(row=total_row, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
    
    # Автоподбор ширины колонок
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем в память
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Формируем имя файла на основе фильтров
    filename_parts = ['work_reports']
    if date_filter:
        filename_parts.append(date_filter)
    if project_select:
        from projects.models import Project
        try:
            project = Project.objects.get(id=project_select)
            filename_parts.append(project.slug or project.name.replace(' ', '_'))
        except Project.DoesNotExist:
            pass
    if project_status:
        filename_parts.append(project_status)
    
    filename = '_'.join(filename_parts) + '.xlsx'
    
    # Возвращаем файл
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response